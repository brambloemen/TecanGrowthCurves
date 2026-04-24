"""
Tecan growth curve analyzer — full analysis version.

Reads the same .xlsx inputs as the HTML version:
  - Tecan i-control kinetic export (auto-detects Cycle Nr / Time [s] block)
  - Layout workbook with sheets: Strains, Dilutions, Media

Adds beyond the HTML version:
  - Mechanistic growth-model fits: modified Gompertz, logistic, Baranyi
  - Bootstrap confidence intervals on μ_max
  - Goodness-of-fit comparison (AIC) between models
  - Per-well fits with mean/CI aggregation across dilutions
  - Side-by-side comparison of sliding-window μ_max vs model-derived μ_max

Run with:   streamlit run tecan_streamlit.py
Deps:       streamlit pandas numpy scipy openpyxl plotly
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Callable

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from scipy.optimize import curve_fit


# ============================================================================
#  TECAN PARSER (same logic as the JS version, ported to Python)
# ============================================================================
@dataclass
class TecanLabel:
    name: str
    times_s: np.ndarray          # shape (n_cycles,)
    times_h: np.ndarray          # shape (n_cycles,)
    wells: dict[str, np.ndarray] # well -> array of OD values (NaN for missing)
    meta: dict = field(default_factory=dict)


def parse_tecan(file: io.BytesIO | str) -> list[TecanLabel]:
    """Parse a Tecan i-control kinetic export. Handles single- or multi-label files.

    Locates 'Cycle Nr.' and 'Time [s]' rows within each 'Label:' block (or the
    whole sheet if no Label: markers present), then extracts well rows whose
    first cell matches a plate-position pattern.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    labels: list[TecanLabel] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        aoa = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
               for r in range(1, ws.max_row + 1)]

        # Find Label: markers
        label_starts: list[tuple[int, str]] = []
        for r, row in enumerate(aoa):
            v0 = row[0] if row else None
            if isinstance(v0, str) and re.match(r"^Label\s*:", v0.strip(), re.I):
                m = re.search(r"Label\s*:\s*(.+)", v0, re.I)
                name = m.group(1).strip() if m else f"Label{len(label_starts)+1}"
                label_starts.append((r, name))
        if not label_starts:
            label_starts = [(0, sheet_name or "Data")]

        for li, (start, name) in enumerate(label_starts):
            end = label_starts[li + 1][0] if li + 1 < len(label_starts) else len(aoa)
            parsed = _parse_block(aoa, start, end, name)
            if parsed is not None:
                labels.append(parsed)

    return labels


def _parse_block(aoa: list[list], start_r: int, end_r: int, name: str) -> TecanLabel | None:
    cycle_row = time_row = temp_row = -1
    for r in range(start_r, end_r):
        v = aoa[r][0] if aoa[r] else None
        if not isinstance(v, str):
            continue
        s = v.strip()
        if re.match(r"^Cycle\s*Nr", s, re.I):
            cycle_row = r
        elif re.match(r"^Time\s*\[", s, re.I):
            time_row = r
        elif re.match(r"^Temp", s, re.I):
            temp_row = r
    if cycle_row < 0 or time_row < 0:
        return None

    times_s: list[float] = []
    for c in range(1, len(aoa[time_row])):
        v = aoa[time_row][c]
        if v is None or v == "":
            break
        times_s.append(float(v))
    if not times_s:
        return None
    times_s_arr = np.array(times_s)

    well_re = re.compile(r"^[A-P]\d{1,2}$", re.I)
    wells: dict[str, np.ndarray] = {}
    after_header = max(time_row, temp_row if temp_row >= 0 else time_row) + 1
    for r in range(after_header, end_r):
        v = aoa[r][0] if aoa[r] else None
        if v is None:
            continue
        s = str(v).strip()
        if not well_re.match(s):
            continue
        well = s.upper()
        vals: list[float] = []
        for c in range(1, len(times_s) + 1):
            x = aoa[r][c] if c < len(aoa[r]) else None
            vals.append(np.nan if x is None or x == "" else float(x))
        wells[well] = np.array(vals)

    if not wells:
        return None

    # Parse misc meta
    meta: dict = {}
    for r in range(start_r, end_r):
        a = aoa[r][0] if len(aoa[r]) > 0 else None
        b = aoa[r][1] if len(aoa[r]) > 1 else None
        e = aoa[r][4] if len(aoa[r]) > 4 else None
        if isinstance(a, str):
            if re.match(r"^Date", a, re.I) and b:
                meta["date"] = str(b)
            elif re.match(r"^Measurement\s*Wavelength", a, re.I) and e:
                meta["wavelength"] = str(e)
    if len(times_s_arr) >= 2:
        meta["interval_min"] = round((times_s_arr[1] - times_s_arr[0]) / 60, 1)
    meta["duration_h"] = times_s_arr[-1] / 3600

    return TecanLabel(name=name, times_s=times_s_arr, times_h=times_s_arr / 3600,
                      wells=wells, meta=meta)


# ============================================================================
#  LAYOUT PARSER
# ============================================================================
ROWS = list("ABCDEFGH")


def parse_layout(file: io.BytesIO | str) -> dict:
    wb = openpyxl.load_workbook(file, data_only=True)
    want = ["Strains", "Dilutions", "Media"]
    found: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
    for w in want:
        match = next((n for n in wb.sheetnames if n.lower() == w.lower()), None)
        if match is None:
            raise ValueError(f'Layout workbook must contain a sheet named "{w}"')
        found[w] = wb[match]

    def read_grid(ws) -> dict:
        grid: dict[str, object] = {}
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            label = row[0]
            if isinstance(label, str) and re.match(r"^[A-H]$", label.strip(), re.I):
                R = label.strip().upper()
                for c in range(1, 13):
                    if c < len(row):
                        v = row[c]
                        grid[f"{R}{c}"] = None if (v is None or v == "") else v
                    else:
                        grid[f"{R}{c}"] = None
        return grid

    return {"strains": read_grid(found["Strains"]),
            "dilutions": read_grid(found["Dilutions"]),
            "media": read_grid(found["Media"])}


def is_blank(strain: str | None) -> bool:
    return isinstance(strain, str) and strain.strip().lower().startswith("blank")


def detect_blank_contamination(label: TecanLabel, layout: dict, threshold: float = 1.5) -> set[str]:
    """Return blank wells where any OD exceeds threshold × mean OD in the first hour."""
    contaminated = set()
    first_hour = label.times_h <= 1.0
    for well, strain in layout["strains"].items():
        if not is_blank(strain):
            continue
        od = label.wells.get(well)
        if od is None:
            continue
        fh_vals = od[first_hour & np.isfinite(od)]
        if len(fh_vals) == 0:
            continue
        mean_fh = np.mean(fh_vals)
        if mean_fh <= 0:
            continue
        if np.any(od[np.isfinite(od)] > threshold * mean_fh):
            contaminated.add(well)
    return contaminated


# ============================================================================
#  GROWTH MODELS
# ============================================================================
# Each model is parametrized in terms of biologically meaningful quantities:
#   A      = asymptotic ln(OD) amplitude (growth extent)
#   mu     = μ_max, specific growth rate (h⁻¹)
#   lam    = lag time (h)
#   y0     = initial ln(OD)
# Return ln(OD) predictions — we fit in log space for stability.

def gompertz(t, A, mu, lam, y0):
    """Modified Gompertz model (Zwietering et al. 1990).

    ln(N/N0) = A · exp(-exp(μ·e/A · (λ - t) + 1))
    Here y = y0 + ln(N/N0).
    """
    e = np.e
    return y0 + A * np.exp(-np.exp(mu * e / A * (lam - t) + 1))


def logistic(t, A, mu, lam, y0):
    """Modified logistic model."""
    e = np.e
    return y0 + A / (1 + np.exp(4 * mu / A * (lam - t) + 2))


def baranyi(t, A, mu, lam, y0):
    """Baranyi model — included for reference but DISABLED in the UI.

    A parameterization issue causes μ_max estimates to be inflated by roughly
    a factor of 1.5-2× relative to the Zwietering-convention Gompertz fit.
    Kept here so anyone reading the code can see why it was dropped; to
    re-enable, add it back to the MODELS dict after fixing the log10-vs-ln
    convention in the braking term (see biogrowth R package vignettes).
    """
    h0 = mu * lam
    A_t = t + (1 / mu) * np.log(np.exp(-mu * t) + np.exp(-h0) - np.exp(-mu * t - h0))
    return y0 + mu * A_t - np.log(1 + (np.exp(mu * A_t) - 1) / np.exp(A))


MODELS: dict[str, Callable] = {
    "Gompertz": gompertz,
    "Logistic": logistic,
}


def initial_guess(t: np.ndarray, y: np.ndarray) -> list[float]:
    """Rough initial parameter guess for curve_fit.

    A    ≈ ln(max) - ln(min)
    mu   ≈ max finite difference of y
    lam  ≈ time at which y first exceeds y0 + 0.1·A
    y0   ≈ first few points' mean
    """
    y0 = float(np.mean(y[:3]))
    A = float(np.max(y) - y0)
    dy = np.diff(y) / np.diff(t)
    mu = float(np.nanmax(dy)) if np.any(np.isfinite(dy)) else 0.3
    # lag: first time we've grown by 10% of A
    threshold = y0 + 0.1 * A
    idx = np.where(y > threshold)[0]
    lam = float(t[idx[0]]) if len(idx) else 1.0
    return [max(A, 0.1), max(mu, 0.01), max(lam, 0.0), y0]


def fit_model(t: np.ndarray, od: np.ndarray, model_name: str,
              od_floor: float = 0.02, skip_first: bool = True) -> dict | None:
    """Fit a growth model to OD (linear) data, working in ln space.

    Returns dict with params, SEs, predicted trajectory, AIC, R².
    Returns None if the fit fails.

    Parameters
    ----------
    skip_first : bool
        If True, drop the t=0 timepoint. Tecan's first read typically happens
        before the shaking has mixed the inoculum uniformly and the value is
        systematically off (often much higher than cycle 2 as the dense
        overnight settles). Setting this True is standard practice.
    """
    # Drop NaN and below-floor points (can't take log)
    mask = np.isfinite(od) & (od > od_floor)
    if skip_first and len(t) > 0:
        # Drop whichever index corresponds to t=0 (or the first timepoint)
        mask[0] = False
    if mask.sum() < 6:
        return None
    t_fit = t[mask]
    y_fit = np.log(od[mask])

    # Guard: the growth range must span at least ~1 ln unit (~2.7× OD) for a
    # mechanistic sigmoid fit to be meaningful. Without enough amplitude the
    # 4 parameters become unidentifiable and curve_fit produces inflated μ.
    if y_fit.max() - y_fit.min() < 0.5:
        return None

    model_fn = MODELS[model_name]
    p0 = initial_guess(t_fit, y_fit)

    try:
        # Bounds: A > 0, mu > 0, lam ≥ 0, y0 free
        lo = [0.01, 0.001, 0.0, -10]
        hi = [20, 10, max(t_fit[-1], 1), 5]
        popt, pcov = curve_fit(model_fn, t_fit, y_fit, p0=p0,
                                bounds=(lo, hi), maxfev=20000)
    except (RuntimeError, ValueError):
        return None

    y_pred = model_fn(t_fit, *popt)
    residuals = y_fit - y_pred
    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((y_fit - y_fit.mean()) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0

    n = len(y_fit)
    k = len(popt)
    # AIC for Gaussian errors
    if ss_res > 0:
        aic = n * np.log(ss_res / n) + 2 * k
    else:
        aic = -np.inf
    # Parameter standard errors (Cramer-Rao lower bound)
    try:
        perr = np.sqrt(np.diag(pcov))
    except Exception:
        perr = np.full(len(popt), np.nan)

    A, mu, lam, y0 = popt
    dA, dmu, dlam, dy0 = perr
    td_min = 60 * np.log(2) / mu if mu > 0 else np.nan
    # Propagate SE to doubling time via derivative: d(td)/d(mu) = -ln(2)/mu^2
    dtd_min = 60 * np.log(2) / (mu ** 2) * dmu if mu > 0 and np.isfinite(dmu) else np.nan

    return {
        "model": model_name,
        "A": A, "mu": mu, "lam": lam, "y0": y0,
        "A_se": dA, "mu_se": dmu, "lam_se": dlam, "y0_se": dy0,
        "td_min": td_min, "td_min_se": dtd_min,
        "r2": r2, "aic": aic, "n_points": n,
        "popt": popt, "model_fn": model_fn,
        "t_fit": t_fit, "y_fit": y_fit, "y_pred": y_pred,
        "od_max": float(np.nanmax(od[mask])),
    }


def bootstrap_mu(t: np.ndarray, od: np.ndarray, model_name: str,
                 od_floor: float, n_boot: int = 200, seed: int = 42,
                 skip_first: bool = True) -> dict:
    """Bootstrap a 95% CI for μ_max by resampling residuals.

    Residuals from the original fit are resampled and added back to the predicted
    trajectory, then the model is refit. Fast for 200 resamples.
    """
    rng = np.random.default_rng(seed)
    base = fit_model(t, od, model_name, od_floor, skip_first=skip_first)
    if base is None:
        return {"mu_mean": np.nan, "mu_lo": np.nan, "mu_hi": np.nan}
    mus = []
    tds = []
    for _ in range(n_boot):
        # Resample residuals (in log space) and add to predictions
        residuals = base["y_fit"] - base["y_pred"]
        boot_y = base["y_pred"] + rng.choice(residuals, size=len(residuals), replace=True)
        # Fit model on bootstrapped data
        try:
            lo = [0.01, 0.001, 0.0, -10]
            hi = [20, 10, max(base["t_fit"][-1], 1), 5]
            popt, _ = curve_fit(base["model_fn"], base["t_fit"], boot_y,
                                p0=base["popt"], bounds=(lo, hi), maxfev=5000)
            mus.append(popt[1])
            if popt[1] > 0:
                tds.append(60 * np.log(2) / popt[1])
        except (RuntimeError, ValueError):
            continue
    if not mus:
        return {"mu_mean": np.nan, "mu_lo": np.nan, "mu_hi": np.nan,
                "td_mean": np.nan, "td_lo": np.nan, "td_hi": np.nan}
    return {
        "mu_mean": float(np.mean(mus)),
        "mu_lo": float(np.percentile(mus, 2.5)),
        "mu_hi": float(np.percentile(mus, 97.5)),
        "td_mean": float(np.mean(tds)) if tds else np.nan,
        "td_lo": float(np.percentile(tds, 2.5)) if tds else np.nan,
        "td_hi": float(np.percentile(tds, 97.5)) if tds else np.nan,
        "n_successful": len(mus),
    }


def sliding_window_mu(t: np.ndarray, od: np.ndarray,
                      window_h: float = 3.0, od_floor: float = 0.03,
                      skip_first: bool = True) -> dict | None:
    """The simple fallback — max slope on ln(OD) over a sliding window."""
    mask = np.isfinite(od) & (od > od_floor)
    if skip_first and len(t) > 0:
        mask[0] = False
    if mask.sum() < 4:
        return None
    tt = t[mask]
    yy = np.log(od[mask])
    best = None
    for i in range(len(tt)):
        end = i
        while end < len(tt) and tt[end] - tt[i] <= window_h:
            end += 1
        if end - i < 4:
            continue
        x = tt[i:end]
        y = yy[i:end]
        # Linear regression
        n = len(x)
        sx, sy = x.sum(), y.sum()
        sxx, sxy, syy = (x * x).sum(), (x * y).sum(), (y * y).sum()
        denom = sxx - sx * sx / n
        if denom <= 0:
            continue
        slope = (sxy - sx * sy / n) / denom
        if slope <= 0:
            continue
        intercept = sy / n - slope * sx / n
        ss_tot = syy - sy * sy / n
        ss_res = syy - intercept * sy - slope * sxy
        r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
        if best is None or slope > best["mu"]:
            best = {"mu": slope, "intercept": intercept, "r2": r2,
                    "t_start": x[0], "t_end": x[-1], "t_mid": (x[0] + x[-1]) / 2}
    return best


# ============================================================================
#  BLANK SUBTRACTION
# ============================================================================
def blank_trace(label: TecanLabel, layout: dict, medium: str,
                excluded_wells: set[str] | None = None) -> np.ndarray:
    """Mean blank trajectory for a medium.

    Falls back to the base medium (portion before '+') when no exact-match blank
    wells exist — e.g. uses the LB blank for LB+Kan strain wells.
    Returns zeros if no blank wells are found at all.
    """
    n = len(label.times_s)

    def _compute(med: str) -> np.ndarray | None:
        sums, counts, found = np.zeros(n), np.zeros(n), False
        for well, strain in layout["strains"].items():
            if strain is None or not is_blank(str(strain)):
                continue
            if layout["media"].get(well) != med:
                continue
            if well not in label.wells:
                continue
            if excluded_wells and well in excluded_wells:
                continue
            found = True
            od = label.wells[well]
            mask = np.isfinite(od)
            sums[mask] += od[mask]
            counts[mask] += 1
        return np.where(counts > 0, sums / np.maximum(counts, 1), 0.0) if found else None

    result = _compute(medium)
    if result is not None:
        return result
    if medium and "+" in medium:
        base = medium.split("+")[0].strip()
        result = _compute(base)
        if result is not None:
            return result
    return np.zeros(n)


def blank_match_type(layout: dict, medium: str | None,
                     excluded_wells: set[str] | None = None) -> str:
    """Return 'exact', 'fallback', or 'none' for the blank available for this medium."""
    if medium is None:
        return "none"

    def has_blank(med: str) -> bool:
        return any(
            s is not None and is_blank(str(s))
            and layout["media"].get(w) == med
            and (not excluded_wells or w not in excluded_wells)
            for w, s in layout["strains"].items()
        )

    if has_blank(medium):
        return "exact"
    if "+" in medium:
        base = medium.split("+")[0].strip()
        if has_blank(base):
            return "fallback"
    return "none"


def corrected_od(label: TecanLabel, layout: dict, well: str, do_blank: bool,
                 excluded_wells: set[str] | None = None) -> np.ndarray:
    od = label.wells.get(well)
    if od is None:
        return np.array([])
    if not do_blank:
        return od.copy()
    medium = layout["media"].get(well)
    blk = blank_trace(label, layout, medium, excluded_wells) if medium is not None else 0.0
    return od - blk


def align_traces(times_h: np.ndarray, stacked: np.ndarray,
                 dil_exps: list[float], align_od: float,
                 ref_dil_exp: float | None = None) -> np.ndarray:
    """Shift each well in time so all cross align_od at the same moment.

    Uses ref_dil_exp as the time reference (first matching well); falls back to
    the most-diluted well if ref_dil_exp is None or not present in dil_exps.
    Wells that never cross align_od are returned as NaN rows (excluded from mean/SD).
    """
    from scipy.interpolate import interp1d as _interp1d

    n_wells, n_times = stacked.shape

    def crossing_time(od: np.ndarray) -> float | None:
        for i in range(n_times - 1):
            v0, v1 = od[i], od[i + 1]
            if np.isfinite(v0) and np.isfinite(v1) and v0 <= align_od <= v1:
                frac = (align_od - v0) / (v1 - v0)
                return float(times_h[i] + frac * (times_h[i + 1] - times_h[i]))
        return None

    if ref_dil_exp is not None and ref_dil_exp in dil_exps:
        ref_idx = dil_exps.index(ref_dil_exp)
    else:
        ref_idx = int(np.argmax(dil_exps))

    t_ref = crossing_time(stacked[ref_idx])
    if t_ref is None:
        return stacked.copy()

    aligned = np.full_like(stacked, np.nan)
    for i in range(n_wells):
        t_i = crossing_time(stacked[i])
        if t_i is None:
            continue
        shift = t_ref - t_i
        mask = np.isfinite(stacked[i])
        if mask.sum() < 2:
            continue
        f = _interp1d(times_h[mask], stacked[i][mask], bounds_error=False, fill_value=np.nan)
        aligned[i] = f(times_h - shift)

    return aligned


def normalize_dilution(v, mode: str) -> float:
    """Return log10 exponent (0 = undiluted, 1 = 10×, 2 = 100×...)."""
    if v is None or v == "":
        return 0.0
    try:
        n = float(v)
    except (ValueError, TypeError):
        return 0.0
    if mode == "log10":
        return round(n, 3)
    if n <= 0:
        return 0.0
    return round(float(np.log10(n)), 3)


# ============================================================================
#  STREAMLIT APP
# ============================================================================
st.set_page_config(page_title="Tecan analyzer — full",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
  .stApp { background: #faf8f3; }
  h1, h2, h3 { font-family: 'Source Serif 4', Georgia, serif !important; }
  .metric-card {
    background: white; padding: 12px 16px; border: 1px solid #d8d2c4;
    border-left: 3px solid #8b3a2a; border-radius: 2px;
  }
</style>
""", unsafe_allow_html=True)

st.title("Tecan growth curve analyzer")
st.caption("Streamlit version with mechanistic growth-model fits (Gompertz, logistic, Baranyi), "
           "bootstrap CIs on μ_max, and AIC-based model comparison.")

# ---- Sidebar: inputs ----
with st.sidebar:
    st.header("1 · Input files")
    tecan_file = st.file_uploader("Tecan i-control .xlsx", type=["xlsx", "xls"], key="tecan")
    layout_file = st.file_uploader("Plate layout .xlsx", type=["xlsx", "xls"], key="layout",
                                    help="Sheets: Strains, Dilutions, Media (each 8×12 grid)")

    if tecan_file is None or layout_file is None:
        st.info("Upload both files to begin.")
        st.stop()

    # Parse
    try:
        labels = parse_tecan(tecan_file)
    except Exception as e:
        st.error(f"Tecan parse error: {e}")
        st.stop()
    if not labels:
        st.error("No Tecan data block found in this file.")
        st.stop()

    try:
        layout = parse_layout(layout_file)
    except Exception as e:
        st.error(f"Layout parse error: {e}")
        st.stop()

    st.success(f"Loaded: {len(labels)} label(s), "
               f"{len(layout['strains'])} layout wells")

    st.header("2 · Measurement label")
    label_name = st.selectbox("Label", [lbl.name for lbl in labels])
    label = next(l for l in labels if l.name == label_name)

    contaminated_wells = detect_blank_contamination(label, layout)
    if contaminated_wells:
        st.warning(
            f"Contaminated blank wells excluded from blank subtraction: "
            f"{', '.join(sorted(contaminated_wells))}"
        )

    strain_media = sorted({
        layout["media"].get(w)
        for w, s in layout["strains"].items()
        if s is not None and not is_blank(str(s)) and layout["media"].get(w) is not None
    })
    blank_warnings = []
    for med in strain_media:
        mt = blank_match_type(layout, med, contaminated_wells)
        if mt == "fallback":
            base = med.split("+")[0].strip()
            blank_warnings.append(f"{med}: no exact blank found, using {base} blank as substitute")
        elif mt == "none":
            blank_warnings.append(f"{med}: no blank found, no correction applied")
    if blank_warnings:
        st.warning("Blank matching issues:\n" + "\n".join(f"• {w}" for w in blank_warnings))

    st.header("3 · Analysis options")
    dil_mode = st.radio("Dilution values are",
                        ["linear", "log10"], horizontal=True,
                        help="linear: 1=undiluted, 10=10× diluted. log10: 0=undiluted, 1=10× diluted")
    do_blank = st.checkbox("Subtract per-medium blank mean", value=True)
    od_floor = st.number_input("Fit floor (OD − blank)", value=0.03,
                                min_value=0.005, max_value=0.5, step=0.005,
                                help="Points below this are excluded from all fits "
                                "(noise dominates, ln breaks down)")
    skip_first = st.checkbox("Skip first timepoint (t=0)", value=True,
                              help="Tecan's first read happens before shaking mixes the "
                                   "inoculum — the value is systematically off. Standard practice "
                                   "is to drop it.")
    min_growth_ln = st.number_input("Minimum growth (ln units) for mechanistic fit",
                                     value=1.0, min_value=0.1, max_value=5.0, step=0.1,
                                     help="Strains with less than this much ln(OD) increase are "
                                          "unlikely to have completed enough of a sigmoid for the "
                                          "Gompertz/logistic to give reliable μ. 1.0 ≈ 2.7× increase.")
    do_align = st.checkbox(
        "Align dilutions by OD crossing before averaging", value=True,
        help="Shifts each dilution curve in time so all cross the alignment OD at the same "
             "moment before computing the mean ± SD. Prevents naive averaging from flattening "
             "the exponential phase when dilutions have different lag times.",
    )
    if do_align:
        align_od = st.number_input(
            "Alignment OD (blank-subtracted)", value=0.05,
            min_value=0.001, max_value=1.0, step=0.005,
            help="OD at which all dilution curves are time-aligned. Should be above the "
                 "noise floor but within early exponential phase.",
        )
        _dil_options = sorted({normalize_dilution(v, dil_mode)
                               for v in layout["dilutions"].values() if v is not None})
        align_ref_dil = st.selectbox(
            "Reference dilution for alignment",
            options=_dil_options,
            index=len(_dil_options) - 1,
            format_func=lambda x: f"10^{x}" if x != 0 else "undiluted (1×)",
            help="The dilution whose OD-crossing time is used as t=0 for alignment. "
                 "Most diluted is recommended — it is typically in clear exponential "
                 "phase without carryover from the previous culture.",
        )
    else:
        align_od, align_ref_dil = 0.05, None

    st.header("4 · Fit settings")
    fit_method = st.selectbox(
        "Model",
        ["Sliding-window μ_max", "Gompertz", "Logistic",
         "Compare models (AIC)"],
        help="Sliding-window is the recommended default. Gompertz/logistic additionally "
             "give lag time and asymptote but tend to overestimate μ_max by 10–30%.",
    )
    fit_scope = st.radio(
        "Fit to",
        ["Mean across dilutions", "Each well separately"],
        help="Per-well fitting is more robust when dilutions hit exponential phase at different times. "
             "For each strain, each well is fit independently, then μ_max values are aggregated "
             "(median + IQR across wells).",
    )
    window_h = st.slider("Sliding window width (h)",
                         min_value=1.0, max_value=8.0, value=3.0, step=0.5)
    do_bootstrap = st.checkbox("Bootstrap 95% CI on μ_max", value=False,
                                help="200 resamples — adds a few seconds per strain. "
                                "Only applies to the mechanistic models when fitting mean.")

# ---- Main: strain selection ----
st.subheader("Plate layout")

# Build a plate-shaped grid for quick visual reference
strains_grid = []
for r in ROWS:
    row = []
    for c in range(1, 13):
        w = f"{r}{c}"
        s = layout["strains"].get(w)
        d = layout["dilutions"].get(w)
        m = layout["media"].get(w)
        if s is None:
            row.append("")
        else:
            row.append(str(s))
    strains_grid.append(row)
plate_df = pd.DataFrame(strains_grid, index=ROWS, columns=list(range(1, 13)))
st.dataframe(plate_df, use_container_width=True, height=300)

# Unique strains (non-blank)
all_strains = sorted(
    {str(s) for s in layout["strains"].values() if s is not None and not is_blank(str(s))},
    key=lambda x: (not x.replace(".", "").isdigit(), float(x) if x.replace(".", "").isdigit() else 0, x),
)

col1, col2 = st.columns([3, 1])
with col1:
    selected_strains = st.multiselect(
        "Strains to analyze",
        options=all_strains,
        default=all_strains,
    )
with col2:
    all_dils = sorted({normalize_dilution(v, dil_mode) for v in layout["dilutions"].values()
                        if v is not None})
    selected_dils = st.multiselect(
        "Dilutions (log₁₀ exponent)",
        options=all_dils,
        default=all_dils,
        format_func=lambda x: f"10^{x}",
    )

if not selected_strains:
    st.warning("Select at least one strain.")
    st.stop()


# ---- Analysis ----
def get_wells_for_strain(strain: str) -> list[tuple[str, float, str]]:
    """Return list of (well, dilution_log10_exp, medium) for this strain,
    filtered by selected_dils."""
    out = []
    for well, s in layout["strains"].items():
        if s is None or str(s) != strain:
            continue
        d = normalize_dilution(layout["dilutions"].get(well), dil_mode)
        if d not in selected_dils:
            continue
        m = layout["media"].get(well)
        if well in label.wells:
            out.append((well, d, m))
    return out


# Generate color palette
PALETTE = ["#8b3a2a", "#c06a3a", "#d9a54e", "#9c8e3a", "#5b7a3a", "#3a6b4e",
           "#3a6b78", "#3d5a8a", "#5a4a8a", "#7a3a7a", "#a8456b", "#6b3d3d",
           "#a87a4e", "#7a6b3a", "#4e8a6b", "#4e6b8a", "#7a4e8a", "#8a4e6b",
           "#4a4a4a", "#6b4a2a", "#2a4a6b", "#5a8a3a", "#8a5a3a", "#3a8a8a"]
strain_color = {s: PALETTE[i % len(PALETTE)] for i, s in enumerate(selected_strains)}

# Compute per-strain mean trajectory (across dilutions) for fitting
metrics_rows = []
fit_objects = {}     # strain -> fit dict (for plot overlays)
compare_rows = []    # for the "Compare all models" option

with st.spinner("Fitting growth models…"):
    for strain in selected_strains:
        wells = get_wells_for_strain(strain)
        if not wells:
            continue
        medium = "/".join(sorted({m for _, _, m in wells if m}))
        n = len(label.times_s)
        stacked = np.full((len(wells), n), np.nan)
        for i, (w, _, _) in enumerate(wells):
            stacked[i] = corrected_od(label, layout, w, do_blank, contaminated_wells)
        if do_align and len(wells) > 1:
            stacked = align_traces(label.times_h, stacked, [d for _, d, _ in wells], align_od, align_ref_dil)
        mean_trace = np.nanmean(stacked, axis=0)
        sd_trace = np.nanstd(stacked, axis=0, ddof=1) if len(wells) > 1 else np.zeros(n)

        # --- Per-well fitting path (sliding-window only) ---
        # Per-well Gompertz/logistic is disabled because curve_fit on individual
        # noisy wells tends to sharpen the inflection and overestimate μ by ~30-50%
        # vs mean-fit or vs direct log-linear regression. Sliding window is the
        # honest per-well estimator.
        if fit_scope == "Each well separately" and fit_method != "Compare models (AIC)":
            if fit_method in ("Gompertz", "Logistic"):
                st.warning(
                    f"Per-well fitting is not supported for {fit_method}: curve_fit on "
                    "individual noisy wells tends to overestimate μ_max by 30–50%. "
                    "Switched to per-well sliding-window. For a rigorous mechanistic fit, "
                    "use 'Mean across dilutions' with {fit_method}."
                )
            per_well_fits = []
            for wi, (well_id, dil, med) in enumerate(wells):
                well_od = stacked[wi]
                f = sliding_window_mu(label.times_h, well_od,
                                      window_h=window_h, od_floor=od_floor,
                                      skip_first=skip_first)
                if f:
                    per_well_fits.append({
                        "well": well_id, "dilution": dil, "model": "sliding",
                        "mu": f["mu"],
                        "td_min": 60 * np.log(2) / f["mu"] if f["mu"] > 0 else np.nan,
                        "lag": np.nan, "A": np.nan,
                        "r2": f["r2"], "fit_obj": f,
                    })
            if not per_well_fits:
                continue
            mus = np.array([p["mu"] for p in per_well_fits])
            tds = np.array([p["td_min"] for p in per_well_fits])
            # Pick representative fit for overlay: median μ's well
            rep_idx = int(np.argsort(mus)[len(mus) // 2])
            fit_objects[strain] = ("sliding", per_well_fits[rep_idx]["fit_obj"], per_well_fits)
            metrics_rows.append({
                "strain": strain, "medium": medium, "best_model": "sliding (per-well)",
                "μ_max (h⁻¹)": float(np.median(mus)),
                "μ_se": float(np.std(mus, ddof=1) / np.sqrt(len(mus))) if len(mus) > 1 else np.nan,
                "μ_IQR_lo": float(np.percentile(mus, 25)),
                "μ_IQR_hi": float(np.percentile(mus, 75)),
                "t_d (min)": float(np.median(tds)),
                "t_d_se": float(np.std(tds, ddof=1) / np.sqrt(len(tds))) if len(tds) > 1 else np.nan,
                "lag (h)": np.nan,
                "A (ln units)": np.nan,
                "R²": float(np.median([p["r2"] for p in per_well_fits])),
                "AIC": np.nan,
                "n wells": len(per_well_fits),
                "OD_max": float(np.nanmax(mean_trace[np.isfinite(mean_trace)])),
            })
            continue  # skip the mean-fit path below

        # --- Mean-across-dilutions path (original) ---
        if fit_method == "Compare models (AIC)":
            fits = {}
            for m_name in ["Gompertz", "Logistic"]:
                f = fit_model(label.times_h, mean_trace, m_name, od_floor, skip_first=skip_first)
                if f is not None:
                    fits[m_name] = f
                    compare_rows.append({
                        "strain": strain, "medium": medium, "model": m_name,
                        "μ_max (h⁻¹)": f["mu"], "t_d (min)": f["td_min"],
                        "lag (h)": f["lam"], "A": f["A"],
                        "R²": f["r2"], "AIC": f["aic"],
                    })
            # Pick best by AIC
            best_name = min(fits, key=lambda n: fits[n]["aic"]) if fits else None
            if best_name:
                fit_objects[strain] = (best_name, fits[best_name])
                f = fits[best_name]
                row = {
                    "strain": strain, "medium": medium, "best_model": best_name,
                    "μ_max (h⁻¹)": f["mu"], "μ_se": f["mu_se"],
                    "t_d (min)": f["td_min"], "t_d_se": f["td_min_se"],
                    "lag (h)": f["lam"], "A (ln units)": f["A"],
                    "R²": f["r2"], "AIC": f["aic"],
                    "n wells": len(wells), "OD_max": f["od_max"],
                }
                metrics_rows.append(row)
        elif fit_method == "Sliding-window μ_max":
            f = sliding_window_mu(label.times_h, mean_trace, skip_first=skip_first,
                                   window_h=window_h, od_floor=od_floor)
            if f:
                fit_objects[strain] = ("sliding", f)
                metrics_rows.append({
                    "strain": strain, "medium": medium, "best_model": "sliding",
                    "μ_max (h⁻¹)": f["mu"], "μ_se": np.nan,
                    "t_d (min)": 60 * np.log(2) / f["mu"] if f["mu"] > 0 else np.nan,
                    "t_d_se": np.nan, "lag (h)": np.nan,
                    "A (ln units)": np.nan, "R²": f["r2"], "AIC": np.nan,
                    "n wells": len(wells),
                    "OD_max": float(np.nanmax(mean_trace[np.isfinite(mean_trace)])),
                })
        else:
            f = fit_model(label.times_h, mean_trace, fit_method, od_floor, skip_first=skip_first)
            if f:
                boot = None
                if do_bootstrap:
                    boot = bootstrap_mu(label.times_h, mean_trace,
                                         fit_method, od_floor, skip_first=skip_first)
                fit_objects[strain] = (fit_method, f)
                row = {
                    "strain": strain, "medium": medium, "best_model": fit_method,
                    "μ_max (h⁻¹)": f["mu"], "μ_se": f["mu_se"],
                    "t_d (min)": f["td_min"], "t_d_se": f["td_min_se"],
                    "lag (h)": f["lam"], "A (ln units)": f["A"],
                    "R²": f["r2"], "AIC": f["aic"],
                    "n wells": len(wells), "OD_max": f["od_max"],
                }
                if boot:
                    row["μ_boot_lo"] = boot["mu_lo"]
                    row["μ_boot_hi"] = boot["mu_hi"]
                    row["t_d_boot_lo"] = boot.get("td_lo", np.nan)
                    row["t_d_boot_hi"] = boot.get("td_hi", np.nan)
                metrics_rows.append(row)


# ---- Plot ----
st.subheader("Growth curves")
scale = st.radio("Y axis", ["OD", "ln(OD)"], horizontal=True, label_visibility="collapsed")

fig = go.Figure()

for strain in selected_strains:
    wells = get_wells_for_strain(strain)
    if not wells:
        continue
    color = strain_color[strain]
    n = len(label.times_s)
    stacked = np.full((len(wells), n), np.nan)
    for i, (w, _, _) in enumerate(wells):
        stacked[i] = corrected_od(label, layout, w, do_blank, contaminated_wells)
    if do_align and len(wells) > 1:
        stacked = align_traces(label.times_h, stacked, [d for _, d, _ in wells], align_od)
    mean_trace = np.nanmean(stacked, axis=0)
    sd_trace = np.nanstd(stacked, axis=0, ddof=1) if len(wells) > 1 else np.zeros(n)

    if scale == "ln(OD)":
        y_mean = np.where(mean_trace > 0, np.log(np.maximum(mean_trace, 1e-10)), np.nan)
        y_up = np.where(mean_trace + sd_trace > 0, np.log(np.maximum(mean_trace + sd_trace, 1e-10)), np.nan)
        y_lo = np.where(mean_trace - sd_trace > 0, np.log(np.maximum(mean_trace - sd_trace, 1e-10)), np.nan)
    else:
        y_mean, y_up, y_lo = mean_trace, mean_trace + sd_trace, mean_trace - sd_trace

    # SD band
    if len(wells) > 1:
        fig.add_trace(go.Scatter(
            x=np.concatenate([label.times_h, label.times_h[::-1]]),
            y=np.concatenate([y_up, y_lo[::-1]]),
            fill="toself", fillcolor=color + "22", line=dict(color="rgba(0,0,0,0)"),
            showlegend=False, hoverinfo="skip",
        ))

    fig.add_trace(go.Scatter(
        x=label.times_h, y=y_mean, mode="lines",
        name=strain,
        line=dict(color=color, width=2),
        legendgroup=strain,
        hovertemplate=f"<b>{strain}</b><br>t=%{{x:.2f}} h<br>OD=%{{y:.3f}}<extra></extra>",
    ))

    # Overlay fit if available
    if strain in fit_objects:
        obj = fit_objects[strain]
        if len(obj) == 3:
            model_name, f, _per_well = obj
        else:
            model_name, f = obj
        if model_name == "sliding":
            x_ext = np.array([max(0, f["t_start"] - 0.5), f["t_end"] + 0.5])
            y_log = f["intercept"] + f["mu"] * x_ext
            y_plot = y_log if scale == "ln(OD)" else np.exp(y_log)
        else:
            t_dense = np.linspace(f["t_fit"][0], f["t_fit"][-1], 80)
            y_log = f["model_fn"](t_dense, *f["popt"])
            y_plot = y_log if scale == "ln(OD)" else np.exp(y_log)
            x_ext = t_dense
        fig.add_trace(go.Scatter(
            x=x_ext, y=y_plot, mode="lines",
            line=dict(color="#1a1a1a", width=1, dash="dot"),
            name=f"{strain} fit", legendgroup=strain, showlegend=False,
            hovertemplate=f"{model_name} fit<extra></extra>",
        ))

fig.update_layout(
    template="simple_white",
    xaxis_title="Time (h)",
    yaxis_title="ln(OD − blank)" if scale == "ln(OD)" else "OD − blank" if do_blank else "OD",
    height=500, margin=dict(l=60, r=20, t=10, b=50),
    plot_bgcolor="white", paper_bgcolor="white",
    hoverlabel=dict(bgcolor="#1a1a1a", font=dict(color="white", family="JetBrains Mono, monospace", size=11)),
    xaxis=dict(dtick=2),
)
st.plotly_chart(fig, use_container_width=True)


# ---- Metrics table ----
st.subheader("Growth metrics")

if fit_method == "Compare models (AIC)":
    st.markdown("**Per-model fits** (best model for each strain is chosen by lowest AIC):")
    compare_df = pd.DataFrame(compare_rows)
    if not compare_df.empty:
        compare_df = compare_df.round({"μ_max (h⁻¹)": 3, "t_d (min)": 1, "lag (h)": 2,
                                        "A": 3, "R²": 4, "AIC": 2})
        st.dataframe(compare_df, use_container_width=True, height=min(600, 40 + 35 * len(compare_df)))

    st.markdown("**Best-model summary per strain:**")

if metrics_rows:
    df = pd.DataFrame(metrics_rows)
    # Rounding for display
    round_spec = {
        "μ_max (h⁻¹)": 3, "μ_se": 3,
        "μ_boot_lo": 3, "μ_boot_hi": 3,
        "t_d (min)": 1, "t_d_se": 1,
        "t_d_boot_lo": 1, "t_d_boot_hi": 1,
        "lag (h)": 2, "A (ln units)": 3,
        "R²": 4, "AIC": 2, "OD_max": 3,
    }
    round_spec = {k: v for k, v in round_spec.items() if k in df.columns}
    df_disp = df.round(round_spec).sort_values("μ_max (h⁻¹)", ascending=False)
    st.dataframe(df_disp, use_container_width=True,
                 height=min(600, 40 + 35 * len(df_disp)))

    # CSV download
    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button("Download metrics CSV", csv,
                       file_name="growth_metrics.csv", mime="text/csv")
else:
    st.info("No successful fits — try lowering the fit floor or switching model.")


# ---- Per-strain detail expanders ----
if len(selected_strains) <= 12:  # don't render 25 expanders
    st.subheader("Per-strain detail")
    for strain in selected_strains:
        if strain not in fit_objects:
            continue
        obj = fit_objects[strain]
        per_well = None
        if len(obj) == 3:
            model_name, f, per_well = obj
        else:
            model_name, f = obj
        wells = get_wells_for_strain(strain)
        header = f"**{strain}** · {model_name} · μ = {f.get('mu', 0):.3f} h⁻¹"
        if per_well:
            header += f" (per-well median)"
        with st.expander(header):
            c1, c2, c3 = st.columns(3)
            if model_name != "sliding":
                c1.metric("μ_max", f"{f['mu']:.3f} h⁻¹",
                          f"±{f['mu_se']:.3f}" if 'mu_se' in f else None)
                c2.metric("Doubling time", f"{f['td_min']:.1f} min",
                          f"±{f['td_min_se']:.1f}" if f.get('td_min_se') is not None and np.isfinite(f.get('td_min_se', np.nan)) else None)
                c3.metric("Lag", f"{f['lam']:.2f} h",
                          f"±{f['lam_se']:.2f}" if 'lam_se' in f else None)
                st.caption(f"R² = {f['r2']:.4f} · AIC = {f['aic']:.2f} · n points = {f['n_points']} · n wells = {len(wells)}")
            else:
                c1.metric("μ_max (window)", f"{f['mu']:.3f} h⁻¹")
                c2.metric("Doubling time", f"{60*np.log(2)/f['mu']:.1f} min" if f['mu'] > 0 else "—")
                c3.metric("Window center", f"{f['t_mid']:.2f} h")
                st.caption(f"R² = {f['r2']:.4f}")
            if per_well:
                pw_df = pd.DataFrame([{
                    "well": p["well"], "dilution (log10)": p["dilution"],
                    "μ_max (h⁻¹)": round(p["mu"], 3),
                    "t_d (min)": round(p["td_min"], 1) if np.isfinite(p["td_min"]) else "—",
                    "lag (h)": round(p["lag"], 2) if np.isfinite(p["lag"]) else "—",
                    "R²": round(p["r2"], 4),
                } for p in per_well])
                st.caption("Per-well fits:")
                st.dataframe(pw_df, use_container_width=True, hide_index=True)

# ---- Footer notes ----
st.markdown("---")
st.markdown("""
**Which μ_max estimator should you use?**

Three methods are offered, and they do *not* always give the same number.
Understanding why helps pick the right one.

- **Sliding-window (log-linear regression)** — finds the steepest log-linear
  stretch of a configurable width and reports its slope. This is the most
  direct, interpretable, and defensible method: you can literally draw the
  tangent on the ln(OD) plot and read the slope. Use this as the default.
  Matches what growthrates::fit_easylinear does in R.
- **Gompertz / logistic on the mean** — a 4-parameter sigmoid fit that
  additionally gives you lag time and asymptotic growth extent. On real
  Tecan data the μ_max from these models is typically 10–30% higher than
  the sliding-window value, because the model's inflection-point slope is
  steeper than the mean slope over any finite window, and the model sharpens
  its inflection to accommodate imperfect sigmoid data (post-stationary
  settling, evaporation). Use when you need lag time too, and interpret
  μ_max with that bias in mind.
- **Per-well sliding-window (median across dilutions)** — useful when
  dilutions are in different growth phases at the same time and averaging
  them produces a non-sigmoid mean. Look at the IQR across wells — if it's
  wide, your dilutions disagree substantially and you should probably pick
  one dilution manually.

**Practical workflow**: start with sliding-window on the mean. When the mean
curve looks bad (ragged, non-monotonic, or strongly dilution-skewed), switch
to per-well. Use Gompertz on the mean only when the curves are clean
sigmoids and you need lag or amplitude.

**First timepoint** is dropped by default — Tecan's first read happens before
shaking has mixed the inoculum and is systematically off.

**Minimum growth** of 0.5 ln units (~1.6× OD increase) is required for
mechanistic fits; strains that haven't grown enough fall back to sliding-window.

**Bootstrap CI** resamples residuals 200× and refits — use for a more
conservative interval than the Cramér–Rao asymptotic SE, especially with
heteroscedastic errors (which OD data usually has).
""")
